from django.shortcuts import redirect, get_object_or_404
from .models import AuditEntry
from django.shortcuts import render
from django.contrib import messages
from django.core.paginator import Paginator
from django.http import HttpResponse
from django.contrib.auth.decorators import login_required
from django.db.models import Count, Q
import csv
import json
from urllib.parse import urlencode
from io import BytesIO


def _parse_changes(raw):
    """Intenta parsear el campo changes como JSON. Devuelve dict o None."""
    if not raw:
        return None
    try:
        return json.loads(raw)
    except Exception:
        return None


def _filtrar_auditorias(request):
    """Aplica filtros comunes para lista y exportación."""
    qs = AuditEntry.objects.select_related('user')
    app = request.GET.get('app', '').strip()
    modelo = request.GET.get('modelo', '').strip()
    evento = request.GET.get('evento', '').strip()
    usuario = request.GET.get('usuario', '').strip()
    fecha_desde = request.GET.get('fecha_desde', '').strip()
    fecha_hasta = request.GET.get('fecha_hasta', '').strip()
    categoria = request.GET.get('categoria', '').strip()
    object_id = request.GET.get('object_id', '').strip()
    ip = request.GET.get('ip', '').strip()
    buscar = request.GET.get('buscar', '').strip()

    if app:
        qs = qs.filter(app_label__icontains=app)
    if modelo:
        qs = qs.filter(model__icontains=modelo)
    if evento:
        qs = qs.filter(action=evento)
    if usuario:
        if usuario.isdigit():
            qs = qs.filter(user_id=usuario)
        else:
            qs = qs.filter(user__email__icontains=usuario)
    if fecha_desde:
        qs = qs.filter(timestamp__date__gte=fecha_desde)
    if fecha_hasta:
        qs = qs.filter(timestamp__date__lte=fecha_hasta)
    if categoria:
        qs = qs.filter(extra__icontains=f'"category": "{categoria}"')
    if object_id:
        qs = qs.filter(object_id=object_id)
    if ip:
        qs = qs.filter(ip_address__icontains=ip)
    if buscar:
        qs = qs.filter(object_repr__icontains=buscar)
    return qs


@login_required
def ver_auditoria(request, pk):
    auditoria = get_object_or_404(AuditEntry, pk=pk)
    changes_parsed = _parse_changes(auditoria.changes)
    extra_parsed = _parse_changes(auditoria.extra)

    # Navegar anterior/siguiente
    prev_entry = AuditEntry.objects.filter(pk__gt=pk).order_by('pk').first()
    next_entry = AuditEntry.objects.filter(pk__lt=pk).order_by('-pk').first()

    return render(request, 'auditoria/detalle_auditoria.html', {
        'auditoria': auditoria,
        'changes_parsed': changes_parsed,
        'extra_parsed': extra_parsed,
        'prev_entry': prev_entry,
        'next_entry': next_entry,
    })


@login_required
def lista_auditoria(request):
    qs = _filtrar_auditorias(request).order_by('-timestamp')

    # Estadísticas sobre el conjunto filtrado (sin paginar)
    stats = qs.aggregate(
        total=Count('id'),
        creates=Count('id', filter=Q(action='create')),
        updates=Count('id', filter=Q(action='update')),
        deletes=Count('id', filter=Q(action='delete')),
    )

    page_size = 20
    paginator = Paginator(qs, page_size)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    export_qs = request.GET.copy()
    if 'page' in export_qs:
        export_qs.pop('page')
    export_url_params = urlencode(export_qs)

    return render(request, 'auditoria/lista_auditoria.html', {
        'auditorias': page_obj,
        'export_params': export_url_params,
        'stats': stats,
    })


@login_required
def eliminar_auditoria(request, pk):
    auditoria = get_object_or_404(AuditEntry, pk=pk)
    if not request.user.is_staff:
        messages.error(request, 'No tiene permisos para eliminar registros de auditoría.')
        return redirect('lista_auditoria')
    if request.method == 'POST':
        auditoria.delete()
        messages.success(request, 'Registro de auditoría eliminado correctamente.')
        return redirect('lista_auditoria')
    return render(request, 'auditoria/confirmar_eliminacion.html', {'auditoria': auditoria})


@login_required
def exportar_auditoria_csv(request):
    """Exporta la auditoría filtrada a CSV."""
    auditorias = _filtrar_auditorias(request).order_by('-timestamp')
    response = HttpResponse(content_type='text/csv; charset=utf-8')
    response['Content-Disposition'] = 'attachment; filename="auditoria.csv"'

    writer = csv.writer(response)
    writer.writerow([
        'timestamp', 'usuario_email', 'app_label', 'model', 'object_id', 'action',
        'path', 'method', 'ip_address', 'changes', 'extra'
    ])
    for a in auditorias:
        writer.writerow([
            a.timestamp.isoformat(),
            getattr(a.user, 'email', '') if a.user else '',
            a.app_label,
            a.model,
            a.object_id,
            a.action,
            a.path,
            a.method,
            a.ip_address or '',
            (a.changes or '').replace('\r', ' ').replace('\n', ' '),
            (a.extra or '').replace('\r', ' ').replace('\n', ' '),
        ])
    return response


@login_required
def exportar_auditoria_xlsx(request):
    """Exporta la auditoría filtrada a Excel (XLSX)."""
    auditorias = _filtrar_auditorias(request).order_by('-timestamp')
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
    except Exception:
        return exportar_auditoria_csv(request)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Auditoria"

    headers = [
        'Fecha/Hora', 'Usuario', 'App', 'Modelo', 'ID Objeto', 'Acción',
        'Ruta', 'Método', 'IP', 'Cambios', 'Extra'
    ]
    ws.append(headers)

    # Estilo encabezado
    header_fill = PatternFill(start_color='1E40AF', end_color='1E40AF', fill_type='solid')
    for col_idx, _ in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')

    action_colors = {'create': 'D1FAE5', 'update': 'DBEAFE', 'delete': 'FEE2E2'}

    for row_idx, a in enumerate(auditorias, start=2):
        action = a.action or ''
        row_data = [
            a.timestamp.strftime('%d/%m/%Y %H:%M:%S'),
            getattr(a.user, 'email', '') if a.user else '',
            a.app_label, a.model, a.object_id, action,
            a.path, a.method, a.ip_address or '',
            (a.changes or ''), (a.extra or ''),
        ]
        ws.append(row_data)
        fill_color = action_colors.get(action)
        if fill_color:
            fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type='solid')
            for col_idx in range(1, len(headers) + 1):
                ws.cell(row=row_idx, column=col_idx).fill = fill

    col_widths = [20, 28, 14, 16, 10, 10, 40, 8, 16, 50, 40]
    for col_idx, width in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    response = HttpResponse(
        buf.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="auditoria.xlsx"'
    return response
